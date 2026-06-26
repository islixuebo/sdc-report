#!/usr/bin/env python3
"""
SDC数字化需求任务统计报告 — 数据拉取与报告生成
配置文件从 .sdc_report_config.json 读取，不在代码中硬编码
"""
import json
from datetime import datetime
import calendar
import re
import os
import sys
import ssl
import urllib.request
import urllib.parse

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ===== 优先从上层目录读取配置 =====
CONFIG_FILE = None
for candidate in [os.path.join(os.path.dirname(BASE_DIR), '.sdc_report_config.json'),
                  os.path.join(BASE_DIR, '.sdc_report_config.json'),
                  os.path.join(os.getcwd(), '.sdc_report_config.json')]:
    if os.path.isfile(candidate):
        CONFIG_FILE = candidate
        break

if not CONFIG_FILE:
    print("=" * 60)
    print("【首次使用】未检测到配置文件 .sdc_report_config.json")
    print("请在项目目录中创建该文件，内容如下：")
    print("""
{
  "jira_url": "https://your-jira-instance.com/rest/api/2/search",
  "jira_token": "your-jira-api-token",
  "jql_active": "project = SDCDN AND issuetype in (任务, 改进) AND fixVersion = EMPTY",
  "fields_base": "key,summary,status,priority,customfield_10348,customfield_10300,customfield_10302,customfield_10458,created,reporter,name,resolutiondate",
  "feishu": {
    "chat_id": "",
    "user_id": ""
  }
}
    """)
    print("或运行初始化脚本：bash scripts/init_config.sh")
    print("=" * 60)
    sys.exit(1)

with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
    config = json.load(f)

JIRA_URL = config.get('jira_url', '')
JIRA_TOKEN = config.get('jira_token', '')
JQL_ACTIVE = config.get('jql_active', 'project = SDCDN AND issuetype in (任务, 改进) AND fixVersion = EMPTY')
FIELDS_BASE = config.get('fields_base', 'key,summary,status,priority,customfield_10348,customfield_10300,customfield_10302,customfield_10458,created,reporter,name,resolutiondate')

if not JIRA_URL or not JIRA_TOKEN:
    print("错误：配置文件中缺少 jira_url 或 jira_token，请检查 .sdc_report_config.json")
    sys.exit(1)

FEISHU_CHAT_ID = config.get('feishu', {}).get('chat_id', '')
FEISHU_USER_ID = config.get('feishu', {}).get('user_id', '')

# ===== 工具函数 =====
def jira_fetch(jql, fields):
    params = urllib.parse.urlencode({'jql': jql, 'maxResults': '200', 'fields': fields})
    req = urllib.request.Request(f"{JIRA_URL}?{params}", method='GET')
    req.add_header('Authorization', f'Bearer {JIRA_TOKEN}')
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(req, context=ctx) as resp:
        return json.loads(resp.read().decode('utf-8')).get('issues', [])

def fmt_date(val):
    if not val:
        return ""
    if isinstance(val, dict):
        s = val.get('value', '') or ''
    else:
        s = str(val)
    s = s.strip()[:10]
    try:
        datetime.strptime(s, '%Y-%m-%d')
        return s
    except:
        return ""

# ===== 1. 从Jira拉取数据 =====
now = datetime.now()
year, month = now.year, now.month
month_start = f"{year}-{month:02d}-01"
month_end = f"{year}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
week_start_date = now - __import__('datetime').timedelta(days=now.weekday())
week_end_date = week_start_date + __import__('datetime').timedelta(days=6)
week_start = week_start_date.strftime('%Y-%m-%d')
week_end = week_end_date.strftime('%Y-%m-%d')

print("正在从Jira拉取数据...")
issues_active = jira_fetch(JQL_ACTIVE, FIELDS_BASE)
print(f"活跃任务: {len(issues_active)} 条")

JQL_DONE_MONTH = f"project = SDCDN AND issuetype in (任务, 改进) AND resolutiondate >= {month_start}"
issues_done_month = jira_fetch(JQL_DONE_MONTH, FIELDS_BASE)
print(f"当月完结任务: {len(issues_done_month)} 条")

issues_dict = {}
for i in issues_active:
    issues_dict[i['key']] = i
for i in issues_done_month:
    key = i['key']
    if key not in issues_dict:
        issues_dict[key] = i
    else:
        res = i.get('fields', {}).get('resolutiondate', '')
        if res:
            issues_dict[key].setdefault('_override_resdate', res)
issues_all = list(issues_dict.values())
print(f"合并后: {len(issues_all)} 条（含当月完结）")

# ===== 2. 数据处理规则 =====
STATUS_MAP = {
    "需求完善": "补充需求",
    "待排期": "排期中",
    "需求验收中": "开发中",
    "已排期": "待开发",
    "开发中": "开发中",
    "数字化审核": "数字化领导审核",
    "所长审核": "需求方所长审核",
    "暂停跟进": "暂停跟进",
    "已发布": "已发布",
    "评审关闭": "已关闭",
    "不是需求": "不是需求",
    "待审核": "待开发",
}

SYSTEM_ORDER = ['RDM', 'SRDPM', 'MAP', 'OSM', 'TOM', 'JIRA', 'SCA', 'AI', 'SINE', 'SOM', 'Skills', 'Devops', '其他']
PRIORITY_ORDER = ['P0', 'P1', 'P2', 'P3', 'P4']
STATUS_ORDER = ['已发布', '已关闭', '待开发', '开发中', '数字化领导审核', '排期中', '需求方所长审核', '补充需求', '暂停跟进', '不是需求']

def sys_rank(s):
    try:
        return SYSTEM_ORDER.index(s) if s in SYSTEM_ORDER else 999
    except:
        return 999

def status_rank(s):
    try:
        return STATUS_ORDER.index(s) if s in STATUS_ORDER else 999
    except:
        return 999

def priority_rank(p):
    try:
        return PRIORITY_ORDER.index(p) if p in PRIORITY_ORDER else 999
    except:
        return 999

def date_rank(d):
    if not d:
        return 0
    try:
        dt = datetime.strptime(str(d).strip()[:10], '%Y-%m-%d')
        return -dt.timestamp()
    except:
        return 0

def clean_system(val):
    if not val:
        return "其他"
    if isinstance(val, dict):
        s = val.get('value', '') or ''
    else:
        s = str(val)
    s = s.strip()
    s = re.sub(r'^【[^】]+】', '', s)
    s = s.strip()
    if not s:
        return "其他"
    for name in SYSTEM_ORDER[:-1]:
        if s.upper() == name.upper():
            return name
    s_upper = s.upper()
    for name in SYSTEM_ORDER[:-1]:
        if name.upper() in s_upper:
            return name
    return "其他"

def clean_reporter(name):
    if not name:
        return ""
    name = name.strip()
    name = re.sub(r'(?:\s+)(SE|SPM|STM|SEng|SWE|Dev|QA|TL|PL|MGR|ENG|FAE|OME|SQM|HW|AE|TPM|PM|PD|PO|BA|DA|SA|SVP|VP|Director|STL)$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'(?<=[\u4e00-\u9fa5a-zA-Z])(SE|SPM|STM|SEng|SWE|Dev|QA|TL|PL|MGR|ENG|FAE|OME|SQM|HW|AE|TPM|PM|PD|PO|BA|DA|SA|SVP|VP|Director|STL)$', '', name, flags=re.IGNORECASE)
    return name.strip()

tasks = []
status_count_raw = {}

for issue in issues_all:
    fields = issue.get('fields', {})
    raw_status = fields.get('status', {}).get('name', '')
    status_count_raw[raw_status] = status_count_raw.get(raw_status, 0) + 1
    pub_status = STATUS_MAP.get(raw_status, raw_status)
    system = clean_system(fields.get('customfield_10348'))
    reporter = clean_reporter(fields.get('reporter', {}).get('displayName', ''))
    prop_date = fmt_date(fields.get('customfield_10458'))
    if not prop_date:
        created = fields.get('created', '')
        prop_date = fmt_date(created)
    delivery = fields.get('customfield_10300')
    if delivery:
        delivery = str(delivery).strip()[:10]
        delivery = fmt_date(delivery)
    else:
        delivery = ""
    summary = fields.get('summary', '')
    priority = fields.get('priority', {})
    if isinstance(priority, dict):
        priority = priority.get('name', '') or ''
    else:
        priority = str(priority) if priority else ''
    priority = priority.strip()
    tasks.append({
        'system': system,
        'summary': summary,
        'date': prop_date,
        'delivery': delivery,
        'reporter': reporter,
        'status': pub_status,
        'raw_status': raw_status,
        'priority': priority,
    })

print(f"处理后: {len(tasks)} 条任务")

# ===== 3. 三层排序 =====
tasks.sort(key=lambda t: (sys_rank(t['system']), priority_rank(t['priority']), status_rank(t['status']), date_rank(t['date'])))

# ===== 3b. 周期分析 =====
DONE_STATUSES = {'已发布', '已关闭'}

def in_month(date_str):
    return month_start <= date_str <= month_end if date_str else False

def in_week(date_str):
    return week_start <= date_str <= week_end if date_str else False

month_new, month_done = [], []
week_new, week_done = [], []

for issue in issues_all:
    fields = issue.get('fields', {})
    created = fmt_date(fields.get('created', ''))
    resdate_raw = issue.get('_override_resdate') or fields.get('resolutiondate', '')
    resdate = fmt_date(resdate_raw)
    if in_month(created):
        month_new.append(issue)
    if in_month(resdate):
        month_done.append(issue)
    if in_week(created):
        week_new.append(issue)
    if in_week(resdate):
        week_done.append(issue)

print(f"\n周期分析:")
print(f"  当月新增: {len(month_new)} 条 | 当月完结: {len(month_done)} 条")
print(f"  当周新增: {len(week_new)} 条 | 当周完结: {len(week_done)} 条")

def period_tasks(issue_list):
    result = []
    for issue in issue_list:
        fields = issue.get('fields', {})
        raw_status = fields.get('status', {}).get('name', '')
        pub_status = STATUS_MAP.get(raw_status, raw_status)
        system = clean_system(fields.get('customfield_10348'))
        reporter = clean_reporter(fields.get('reporter', {}).get('displayName', ''))
        prop_date = fmt_date(fields.get('customfield_10458')) or fmt_date(fields.get('created', ''))
        delivery = fmt_date(fields.get('customfield_10300', ''))
        priority = fields.get('priority', {})
        if isinstance(priority, dict):
            priority = priority.get('name', '') or ''
        else:
            priority = str(priority) if priority else ''
        priority = priority.strip()
        result.append({
            'key': issue.get('key', ''),
            'system': system,
            'summary': fields.get('summary', ''),
            'date': prop_date,
            'delivery': delivery,
            'reporter': reporter,
            'status': pub_status,
            'raw_status': raw_status,
            'priority': priority,
        })
    return result

month_new_tasks = period_tasks(month_new)
month_done_tasks = period_tasks(month_done)
week_new_tasks = period_tasks(week_new)
week_done_tasks = period_tasks(week_done)

# ===== 4. 统计 =====
from collections import Counter

sys_counts = Counter(t['system'] for t in tasks)
status_counts = Counter(t['status'] for t in tasks)
priority_counts = Counter(t['priority'] if t['priority'] else '(未设)' for t in tasks)

print("\n按系统分布:")
for name in SYSTEM_ORDER:
    if sys_counts.get(name, 0) > 0:
        print(f"  {name}: {sys_counts[name]}")

print("\n按状态分布:")
for name in STATUS_ORDER:
    if status_counts.get(name, 0) > 0:
        print(f"  {name}: {status_counts[name]}")

print("\n按优先级分布:")
for name in PRIORITY_ORDER + ['(未设)']:
    if priority_counts.get(name, 0) > 0:
        print(f"  {name}: {priority_counts[name]}")

# ===== 5. 版本号管理 =====
import shutil
from collections import defaultdict

today_str = datetime.now().strftime('%Y-%m-%d')
HISTORY_DIR = os.path.join(BASE_DIR, '历史报告')
REPORT_BASE = 'SDC数字化需求任务统计报告'
EXTS = ['.md', '.xlsx', '.pptx']

os.makedirs(HISTORY_DIR, exist_ok=True)

date_versions = defaultdict(set)
date_version_files = defaultdict(lambda: defaultdict(list))

for fname in os.listdir(BASE_DIR):
    if not fname.startswith(REPORT_BASE):
        continue
    if any(fname.lower().endswith(ext) for ext in EXTS):
        m = re.match(r'SDC数字化需求任务统计报告_(\d{4}-\d{2}-\d{2})\.V(\d+)\.\w+$', fname, re.IGNORECASE)
        if m:
            date_str = m.group(1)
            version = int(m.group(2))
            if date_str != today_str:
                date_versions[date_str].add(version)
                date_version_files[date_str][version].append(fname)

for date_str in date_versions:
    max_v = max(date_versions[date_str])
    for version, fnames in date_version_files[date_str].items():
        for fname in fnames:
            fpath = os.path.join(BASE_DIR, fname)
            if not os.path.isfile(fpath):
                continue
            if version == max_v:
                shutil.move(fpath, os.path.join(HISTORY_DIR, fname))
                print(f"已归档: {fname} → 历史/")
            else:
                os.remove(fpath)
                print(f"已删除（旧版本）: {fname}")

max_ver = 0
file_prefix_today = f'{REPORT_BASE}_{today_str}'
for fname in os.listdir(BASE_DIR):
    if not fname.startswith(file_prefix_today):
        continue
    m = re.match(r'.*\.V(\d+)\.\w+$', fname, re.IGNORECASE)
    if m:
        max_ver = max(max_ver, int(m.group(1)))

current_ver = max_ver + 1
VER_TAG = f'.V{current_ver:02d}'

for fname in list(os.listdir(BASE_DIR)):
    if not fname.startswith(file_prefix_today):
        continue
    if any(fname.lower().endswith(ext) for ext in EXTS):
        fpath = os.path.join(BASE_DIR, fname)
        if os.path.isfile(fpath) and fname != f'{file_prefix_today}{VER_TAG}.md' and fname != f'{file_prefix_today}{VER_TAG}.xlsx' and fname != f'{file_prefix_today}{VER_TAG}.pptx':
            shutil.move(fpath, os.path.join(HISTORY_DIR, fname))
            print(f"已归档（旧版本）: {fname} → 历史/")

# ===== 6. 生成Markdown报告 =====
md_path = os.path.join(BASE_DIR, f'SDC数字化需求任务统计报告_{today_str}{VER_TAG}.md')

md = []
md.append("# SDC数字化需求任务统计报告\n")
md.append("## 1. 背景说明\n")
md.append(f"- **数据来源**: JIRA Filter #11503 - SDC 数字化需求（未发布）")
md.append(f"- **查询条件**: `project = SDCDN AND issuetype in (任务, 改进) AND fixVersion = EMPTY`")
md.append(f"- **生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
md.append(f"- **任务总数**: {len(tasks)}")
md.append(f"- **项目**: SDC 数字化需求 (SDCDN)\n")

md.append("## 2. 状态对应表\n")
md.append("| JIRA状态 | 公示状态 | 说明 |")
md.append("|---------|---------|------|")
for jira_s, pub_s in STATUS_MAP.items():
    cnt = status_count_raw.get(jira_s, 0)
    if cnt > 0:
        md.append(f"| {jira_s} | {pub_s} | 本报告 {cnt} 个 |")
md.append("")

md.append("## 3. 任务统计总览\n")
md.append("### 按系统分布\n")
md.append("| 系统 | 任务数 | 占比 |")
md.append("|------|-------|------|")
for name in SYSTEM_ORDER:
    cnt = sys_counts.get(name, 0)
    if cnt > 0:
        pct = f"{cnt/len(tasks)*100:.1f}%"
        md.append(f"| {name} | {cnt} | {pct} |")
md.append(f"| **合计** | **{len(tasks)}** | **100%** |\n")

md.append("### 按状态分布\n")
md.append("| 公示状态 | 任务数 | 占比 |")
md.append("|---------|-------|------|")
for name in STATUS_ORDER:
    cnt = status_counts.get(name, 0)
    if cnt > 0:
        pct = f"{cnt/len(tasks)*100:.1f}%"
        md.append(f"| {name} | {cnt} | {pct} |")
md.append(f"| **合计** | **{len(tasks)}** | **100%** |\n")

md.append("### 原始JIRA状态分布（验证用）\n")
md.append("| JIRA状态 | 任务数 |")
md.append("|---------|-------|")
for name in sorted(status_count_raw.keys(), key=lambda x: -status_count_raw[x]):
    md.append(f"| {name} | {status_count_raw[name]} |")
md.append("")

md.append("## 4. 详细任务清单\n")
md.append("> 排序规则：系统优先级 → 状态优先级 → 创建日期倒序\n")

current_sys = None
for t in tasks:
    if t['system'] != current_sys:
        current_sys = t['system']
        sys_tasks = [x for x in tasks if x['system'] == current_sys]
        md.append(f"\n### {current_sys}系统 ({len(sys_tasks)}个任务)\n")
        md.append("| 系统 | 概要 | 创建日期 | 交付日 | 报告人 | 状态 |")
        md.append("|------|------|----------|--------|--------|------|")
    md.append(f"| {t['system']} | {t['summary']} | {t['date']} | {t['delivery']} | {t['reporter']} | {t['status']} |")

md.append("")
md.append("## 5. 周期分析\n")
md.append(f"> 周期定义：当月 = {month_start} ~ {month_end} | 当周 = {week_start}（周一）~ {week_end}（周日）\n")

md.append("### 当月统计\n")
md.append("| 指标 | 数量 |")
md.append("|------|------|")
md.append(f"| 当月新增 | {len(month_new_tasks)} |")
md.append(f"| 当月完结 | {len(month_done_tasks)} |")
md.append(f"| 当前进行中 | {len(tasks)} |\n")

md.append("### 本周统计\n")
md.append("| 指标 | 数量 |")
md.append("|------|------|")
md.append(f"| 本周新增 | {len(week_new_tasks)} |")
md.append(f"| 本周完结 | {len(week_done_tasks)} |\n")

def md_period_table_block(md, title, task_list):
    md.append(f"\n### {title}\n")
    md.append("| 系统 | 概要 | 创建日期 | 交付日 | 报告人 | 状态 |")
    md.append("|------|------|----------|--------|--------|------|")
    if not task_list:
        md.append("| - | （无数据） | - | - | - | - |")
    else:
        for t in task_list:
            md.append(f"| {t['system']} | {t['summary']} | {t['date']} | {t['delivery']} | {t['reporter']} | {t['status']} |")
        by_sys = Counter(t['system'] for t in task_list)
        parts = [f"{k}={v}条" for k, v in sorted(by_sys.items()) if v > 0]
        md.append(f"\n_系统分布：{', '.join(parts)}）")

md_period_table_block(md, "当月新增任务清单", month_new_tasks)
md_period_table_block(md, "当月完结任务清单", month_done_tasks)
md_period_table_block(md, "本周新增任务清单", week_new_tasks)
md_period_table_block(md, "本周完结任务清单", week_done_tasks)

md.append("")
md.append("## 6. 优先级分析\n")
md.append("### 按优先级分布\n")
md.append("| 优先级 | 任务数 | 占比 |")
md.append("|--------|-------|------|")
for name in PRIORITY_ORDER:
    cnt = priority_counts.get(name, 0)
    if cnt > 0:
        pct = f"{cnt/len(tasks)*100:.1f}%"
        md.append(f"| {name} | {cnt} | {pct} |")
cnt_none = priority_counts.get('(未设)', 0)
if cnt_none > 0:
    pct = f"{cnt_none/len(tasks)*100:.1f}%"
    md.append(f"| (未设) | {cnt_none} | {pct} |")
md.append(f"| **合计** | **{len(tasks)}** | **100%** |\n")

md.append("### 按优先级任务清单\n")
current_priority = None
for t in tasks:
    p = t['priority'] if t['priority'] else '(未设)'
    if p != current_priority:
        current_priority = p
        pri_tasks = [x for x in tasks if (x['priority'] if x['priority'] else '(未设)') == current_priority]
        md.append(f"\n#### {current_priority}级 ({len(pri_tasks)}个任务)\n")
        md.append("| 系统 | 概要 | 创建日期 | 交付日 | 报告人 | 状态 |")
        md.append("|------|------|----------|--------|--------|------|")
    md.append(f"| {t['system']} | {t['summary']} | {t['date']} | {t['delivery']} | {t['reporter']} | {t['status']} |")
md.append("")

with open(md_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(md))
print(f"\nMarkdown已生成: {md_path}")

# ===== 7. 生成Excel报告 =====
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()

header_font = Font(bold=True)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12)

# 统计总览 Sheet
ws_overview = wb.active
ws_overview.title = '统计总览'

ws_overview.cell(1, 1, f'SDC数字化需求任务统计报告 - {today_str}')
ws_overview.cell(1, 1).font = title_font
ws_overview.merge_cells('A1:D1')

row = 3
ws_overview.cell(row, 1, '按系统分布')
ws_overview.cell(row, 1).font = section_font
row += 1
for c, h in enumerate(['系统', '任务数', '占比', ''], 1):
    cell = ws_overview.cell(row, c, h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
row += 1
for name in SYSTEM_ORDER:
    cnt = sys_counts.get(name, 0)
    if cnt > 0:
        pct = f"{cnt/len(tasks)*100:.1f}%"
        ws_overview.cell(row, 1, name).border = thin_border
        ws_overview.cell(row, 2, cnt).border = thin_border
        ws_overview.cell(row, 2).alignment = Alignment(horizontal='center')
        ws_overview.cell(row, 3, pct).border = thin_border
        ws_overview.cell(row, 3).alignment = Alignment(horizontal='center')
        row += 1
ws_overview.cell(row, 1, '合计').font = Font(bold=True)
ws_overview.cell(row, 1).border = thin_border
ws_overview.cell(row, 2, len(tasks)).font = Font(bold=True)
ws_overview.cell(row, 2).border = thin_border
ws_overview.cell(row, 2).alignment = Alignment(horizontal='center')
ws_overview.cell(row, 3, '100%').font = Font(bold=True)
ws_overview.cell(row, 3).border = thin_border
ws_overview.cell(row, 3).alignment = Alignment(horizontal='center')
row += 2

ws_overview.cell(row, 1, '按状态分布')
ws_overview.cell(row, 1).font = section_font
row += 1
for c, h in enumerate(['公示状态', '任务数', '占比', ''], 1):
    cell = ws_overview.cell(row, c, h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
row += 1
for name in STATUS_ORDER:
    cnt = status_counts.get(name, 0)
    if cnt > 0:
        pct = f"{cnt/len(tasks)*100:.1f}%"
        ws_overview.cell(row, 1, name).border = thin_border
        ws_overview.cell(row, 2, cnt).border = thin_border
        ws_overview.cell(row, 2).alignment = Alignment(horizontal='center')
        ws_overview.cell(row, 3, pct).border = thin_border
        ws_overview.cell(row, 3).alignment = Alignment(horizontal='center')
        row += 1

ws_overview.cell(row, 1, '合计').font = Font(bold=True)
ws_overview.cell(row, 1).border = thin_border
ws_overview.cell(row, 2, len(tasks)).font = Font(bold=True)
ws_overview.cell(row, 2).border = thin_border
ws_overview.cell(row, 2).alignment = Alignment(horizontal='center')
ws_overview.cell(row, 3, '100%').font = Font(bold=True)
ws_overview.cell(row, 3).border = thin_border

ws_overview.column_dimensions['A'].width = 18
ws_overview.column_dimensions['B'].width = 12
ws_overview.column_dimensions['C'].width = 12

# 全部任务 Sheet
ws_all = wb.create_sheet('全部任务')
headers = ['系统', '概要', '创建日期', '交付日', '报告人', '状态']
for c, h in enumerate(headers, 1):
    cell = ws_all.cell(1, c, h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for r_idx, t in enumerate(tasks, 2):
    vals = [t['system'], t['summary'], t['date'], t['delivery'], t['reporter'], t['status']]
    for c_idx, val in enumerate(vals, 1):
        cell = ws_all.cell(r_idx, c_idx, val)
        cell.border = thin_border
        if c_idx in (1, 3, 4, 5, 6):
            cell.alignment = Alignment(horizontal='center')
        if c_idx == 6:
            status_colors = {
                '排期中': '3B82F6', '需求方所长审核': 'F59E0B', '补充需求': 'EF4444',
                '暂停跟进': '94A3B8', '待开发': '8B5CF6', '已发布': '10B981', '已关闭': '64748B',
            }
            clr = status_colors.get(val, '')
            if clr:
                cell.font = Font(color=clr, bold=True)

ws_all.column_dimensions['A'].width = 10
ws_all.column_dimensions['B'].width = 60
ws_all.column_dimensions['C'].width = 14
ws_all.column_dimensions['D'].width = 24
ws_all.column_dimensions['E'].width = 10
ws_all.column_dimensions['F'].width = 16

# 按系统分Sheet
for sys_name in SYSTEM_ORDER:
    sys_tasks_list = [t for t in tasks if t['system'] == sys_name]
    if not sys_tasks_list:
        continue
    ws_sys = wb.create_sheet(sys_name)
    for c, h in enumerate(headers, 1):
        cell = ws_sys.cell(1, c, h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    for r_idx, t in enumerate(sys_tasks_list, 2):
        vals = [t['system'], t['summary'], t['date'], t['delivery'], t['reporter'], t['status']]
        for c_idx, val in enumerate(vals, 1):
            cell = ws_sys.cell(r_idx, c_idx, val)
            cell.border = thin_border
            if c_idx in (1, 3, 4, 5, 6):
                cell.alignment = Alignment(horizontal='center')
    ws_sys.column_dimensions['A'].width = 10
    ws_sys.column_dimensions['B'].width = 60
    ws_sys.column_dimensions['C'].width = 14
    ws_sys.column_dimensions['D'].width = 24
    ws_sys.column_dimensions['E'].width = 10
    ws_sys.column_dimensions['F'].width = 16

wb.move_sheet('全部任务', offset=-len(wb.sheetnames)+2)

# 周期分析 Excel Sheet
def make_period_sheet(wb, sheet_name, task_list):
    ws = wb.create_sheet(sheet_name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    for r_idx, t in enumerate(task_list, 2):
        vals = [t['system'], t['summary'], t['date'], t['delivery'], t['reporter'], t['status']]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.border = thin_border
            if c_idx in (1, 3, 4, 5, 6):
                cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 16
    return ws

make_period_sheet(wb, '当月新增', month_new_tasks)
make_period_sheet(wb, '当月完结', month_done_tasks)
make_period_sheet(wb, '本周新增', week_new_tasks)
make_period_sheet(wb, '本周完结', week_done_tasks)

# 优先级 Excel Sheet
ws_pri = wb.create_sheet('按优先级')
pri_headers = ['优先级', '系统', '概要', '创建日期', '交付日', '报告人', '状态']
pri_header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
for c, h in enumerate(pri_headers, 1):
    cell = ws_pri.cell(1, c, h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = pri_header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

pri_colors = {
    'P0': 'DC2626', 'P1': 'EA580C', 'P2': '2563EB',
    'P3': '16A34A', 'P4': '6B7280', '(未设)': '9CA3AF',
}
row_idx = 2
for p_name in PRIORITY_ORDER + ['(未设)']:
    cnt = priority_counts.get(p_name, 0)
    if cnt == 0:
        continue
    pri_tasks_list = [t for t in tasks if (t['priority'] if t['priority'] else '(未设)') == p_name]
    for t in pri_tasks_list:
        vals = [p_name, t['system'], t['summary'], t['date'], t['delivery'], t['reporter'], t['status']]
        for c_idx, val in enumerate(vals, 1):
            cell = ws_pri.cell(row_idx, c_idx, val)
            cell.border = thin_border
            if c_idx in (1, 2, 3, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal='center')
            if c_idx == 1:
                cell.font = Font(color=pri_colors.get(p_name, '000000'), bold=True)
        row_idx += 1

ws_pri.column_dimensions['A'].width = 10
ws_pri.column_dimensions['B'].width = 10
ws_pri.column_dimensions['C'].width = 55
ws_pri.column_dimensions['D'].width = 14
ws_pri.column_dimensions['E'].width = 24
ws_pri.column_dimensions['F'].width = 10
ws_pri.column_dimensions['G'].width = 16

xlsx_path = os.path.join(BASE_DIR, f'SDC数字化需求任务统计报告_{today_str}{VER_TAG}.xlsx')
wb.save(xlsx_path)
print(f"Excel已生成: {xlsx_path}")

print(f"\n===== 完成 =====")
print(f"任务总数: {len(tasks)}")
print(f"版本号: V{current_ver:02d}")
print(f"Markdown: {os.path.basename(md_path)}")
print(f"Excel: {os.path.basename(xlsx_path)}")
print(f"VER_TAG={VER_TAG}")
print(f"TODAY={today_str}")

# 写版本信息供PPT脚本读取
ver_info_path = os.path.join(BASE_DIR, '.ver_info.tmp')
with open(ver_info_path, 'w', encoding='utf-8') as f:
    f.write(f"{today_str}\n{VER_TAG}\n")

# 导出PPT所需数据
ppt_sys_stats = [{'name': name, 'count': sys_counts.get(name, 0), 'pct': f"{sys_counts.get(name, 0)/len(tasks)*100:.1f}%"} for name in SYSTEM_ORDER if sys_counts.get(name, 0) > 0]
ppt_status_stats = [{'name': name, 'count': status_counts.get(name, 0), 'pct': f"{status_counts.get(name, 0)/len(tasks)*100:.1f}%"} for name in STATUS_ORDER if status_counts.get(name, 0) > 0]
ppt_priority_stats = [{'name': name, 'count': priority_counts.get(name, 0), 'pct': f"{priority_counts.get(name, 0)/len(tasks)*100:.1f}%"} for name in PRIORITY_ORDER + ['(未设)'] if priority_counts.get(name, 0) > 0]

def ppt_tasks(task_list):
    return [[t['system'], t['summary'], t['date'], t['delivery'], t['reporter'], t['status']] for t in task_list]

def ppt_tasks_with_priority(task_list):
    return [[t['priority'] or '(未设)', t['system'], t['summary'], t['date'], t['delivery'], t['reporter'], t['status']] for t in task_list]

ppt_data = {
    'today': today_str,
    'ver_tag': VER_TAG,
    'file_prefix': file_prefix_today,
    'total_tasks': len(tasks),
    'sys_stats': ppt_sys_stats,
    'status_stats': ppt_status_stats,
    'priority_stats': ppt_priority_stats,
    'all_tasks': [[t['system'], t['summary'], t['date'], t['delivery'], t['reporter'], t['status']] for t in tasks],
    'priority_tasks': ppt_tasks_with_priority(tasks),
    'month_range': f"{month_start} ~ {month_end}",
    'week_range': f"{week_start} ~ {week_end}",
    'month_new': {'count': len(month_new_tasks), 'tasks': ppt_tasks(month_new_tasks)},
    'month_done': {'count': len(month_done_tasks), 'tasks': ppt_tasks(month_done_tasks)},
    'week_new': {'count': len(week_new_tasks), 'tasks': ppt_tasks(week_new_tasks)},
    'week_done': {'count': len(week_done_tasks), 'tasks': ppt_tasks(week_done_tasks)},
}
ppt_data_path = os.path.join(BASE_DIR, '.ppt_data.json')
with open(ppt_data_path, 'w', encoding='utf-8') as f:
    json.dump(ppt_data, f, ensure_ascii=False)